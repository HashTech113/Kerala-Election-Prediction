"""Build the Kerala constituency-level actual historical trend / swing dataset.

Reads the real fields available in:
    backend/data_files/kerala_constituency_master_2026.csv
    backend/data_files/kerala_assembly_2026.csv

Writes:
    backend/data_files/kerala_actual_historical_trend_swing_constituencies.csv
    backend/data_files/kerala_actual_historical_trend_swing_constituencies.xlsx

This script intentionally:
    * does NOT use proj_2026_* columns or any 2026 projection fields
    * does NOT invent 2011 Assembly / 2014 LS values (they are not in the dataset)
    * does NOT overwrite historical election CSVs
    * marks unavailable fields with explicit sentinels rather than blanks

Run:
    python backend/build_historical_trend_swing.py
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data_files"
MASTER_FILE = DATA_DIR / "kerala_constituency_master_2026.csv"
ASSEMBLY_FILE = DATA_DIR / "kerala_assembly_2026.csv"
OUT_CSV = DATA_DIR / "kerala_actual_historical_trend_swing_constituencies.csv"
OUT_XLSX = DATA_DIR / "kerala_actual_historical_trend_swing_constituencies.xlsx"

UNAVAILABLE_VALUE = "Not available in uploaded dataset"
UNAVAILABLE_TREND = "Cannot calculate from uploaded dataset"

OUTPUT_COLUMNS = [
    "ac_no",
    "constituency",
    "district",
    "region_5way",
    "reservation",
    "2011 Assembly Actual Winner",
    "2014 Lok Sabha AC/Segment Actual Winner",
    "Historical Projection [2011-2014]",
    "2016 Assembly Actual Winner",
    "2021 Assembly Actual Winner",
    "Long-Term Trend [2014-2021]",
    "2021 Assembly Actual Winner For Swing",
    "2024 Lok Sabha Actual Winner",
    "2024 UDF Vote Share (%)",
    "2024 LDF Vote Share (%)",
    "2024 NDA Vote Share (%)",
    "Recent Swing [2021-2024]",
]

NOTES_TEXT = (
    "This workbook contains actual historical fields available in the uploaded "
    "Kerala dataset only.\n"
    "It excludes 2026 projection fields.\n"
    "2011 Assembly and 2014 Lok Sabha assembly-segment actual winner fields are "
    "not present in the uploaded dataset.\n"
    "Those values are marked as unavailable instead of being guessed.\n"
    "Long-Term Trend [2014-2021] is built using winner_2016 and winner_2021.\n"
    "Recent Swing [2021-2024] is built using winner_2021 and ls2024_winner."
)


def _read_csv(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        return list(csv.DictReader(fp))


def _norm(name: str) -> str:
    """Normalize a constituency name for cross-file matching."""
    return (name or "").strip().lower()


def _trend_label(prev: str, curr: str) -> str:
    prev = (prev or "").strip()
    curr = (curr or "").strip()
    if not prev or not curr:
        return UNAVAILABLE_TREND
    if prev == curr:
        return f"Stable {prev}"
    return f"Shift {prev} to {curr}"


def _pct(value: str) -> str:
    """Convert a fraction string ('0.4515') to a percent rounded to 2 dp ('45.15')."""
    try:
        return f"{round(float(value) * 100.0, 2)}"
    except (TypeError, ValueError):
        return ""


def main() -> None:
    print(f"[build_hts] Loading {MASTER_FILE.name} ...")
    master_rows = _read_csv(MASTER_FILE)
    print(f"[build_hts] master rows: {len(master_rows)}")

    print(f"[build_hts] Loading {ASSEMBLY_FILE.name} ...")
    assembly_rows = _read_csv(ASSEMBLY_FILE)
    print(f"[build_hts] assembly rows: {len(assembly_rows)}")

    # Index assembly rows by normalized constituency name
    assembly_by_name: dict[str, dict] = {}
    for r in assembly_rows:
        key = _norm(r.get("constituency"))
        if key in assembly_by_name:
            print(f"[build_hts] WARNING duplicate constituency in assembly file: {r.get('constituency')!r}")
        assembly_by_name[key] = r

    # Build output rows in master order (so ac_no is sequential 1..140)
    output_rows: list[dict] = []
    unmatched_in_assembly: list[str] = []
    seen_constituencies: set[str] = set()

    for m in master_rows:
        ac_name = (m.get("ac_name") or "").strip()
        ac_no = m.get("ac_no") or ""
        district = (m.get("district") or "").strip()
        region = (m.get("region_5way") or "").strip()
        reservation = (m.get("reservation") or "").strip()

        key = _norm(ac_name)
        if key in seen_constituencies:
            print(f"[build_hts] WARNING duplicate constituency in master: {ac_name!r}")
        seen_constituencies.add(key)

        a = assembly_by_name.get(key)
        if a is None:
            unmatched_in_assembly.append(ac_name)
            winner_2016 = winner_2021 = ls2024_winner = ""
            udf_pct = ldf_pct = nda_pct = ""
        else:
            winner_2016 = (a.get("winner_2016") or "").strip()
            winner_2021 = (a.get("winner_2021") or "").strip()
            ls2024_winner = (a.get("ls2024_winner") or "").strip()
            udf_pct = _pct(a.get("ls2024_udf_pct"))
            ldf_pct = _pct(a.get("ls2024_ldf_pct"))
            nda_pct = _pct(a.get("ls2024_nda_pct"))

        long_term = _trend_label(winner_2016, winner_2021)
        recent_swing = _trend_label(winner_2021, ls2024_winner)

        output_rows.append({
            "ac_no": ac_no,
            "constituency": ac_name,
            "district": district,
            "region_5way": region,
            "reservation": reservation,
            "2011 Assembly Actual Winner": UNAVAILABLE_VALUE,
            "2014 Lok Sabha AC/Segment Actual Winner": UNAVAILABLE_VALUE,
            "Historical Projection [2011-2014]": UNAVAILABLE_TREND,
            "2016 Assembly Actual Winner": winner_2016,
            "2021 Assembly Actual Winner": winner_2021,
            "Long-Term Trend [2014-2021]": long_term,
            "2021 Assembly Actual Winner For Swing": winner_2021,
            "2024 Lok Sabha Actual Winner": ls2024_winner,
            "2024 UDF Vote Share (%)": udf_pct,
            "2024 LDF Vote Share (%)": ldf_pct,
            "2024 NDA Vote Share (%)": nda_pct,
            "Recent Swing [2021-2024]": recent_swing,
        })

    # ---------------- validation ----------------
    errors: list[str] = []
    warnings: list[str] = []

    if len(output_rows) != 140:
        errors.append(f"row count {len(output_rows)} != 140")

    if unmatched_in_assembly:
        errors.append(
            f"{len(unmatched_in_assembly)} master constituencies not found in assembly file: "
            f"{unmatched_in_assembly[:5]}{'...' if len(unmatched_in_assembly) > 5 else ''}"
        )

    seen = set()
    dups = []
    for r in output_rows:
        key = _norm(r["constituency"])
        if key in seen:
            dups.append(r["constituency"])
        seen.add(key)
    if dups:
        errors.append(f"duplicate constituency names: {dups}")

    blank_w16 = sum(1 for r in output_rows if not r["2016 Assembly Actual Winner"])
    blank_w21 = sum(1 for r in output_rows if not r["2021 Assembly Actual Winner"])
    blank_ls24 = sum(1 for r in output_rows if not r["2024 Lok Sabha Actual Winner"])
    if blank_w16:
        errors.append(f"{blank_w16} rows blank in 2016 Assembly Actual Winner")
    if blank_w21:
        errors.append(f"{blank_w21} rows blank in 2021 Assembly Actual Winner")
    if blank_ls24:
        errors.append(f"{blank_ls24} rows blank in 2024 Lok Sabha Actual Winner")

    # numeric vote-share check
    bad_numeric = 0
    for r in output_rows:
        for col in ("2024 UDF Vote Share (%)", "2024 LDF Vote Share (%)", "2024 NDA Vote Share (%)"):
            v = r[col]
            if v == "":
                bad_numeric += 1
                continue
            try:
                float(v)
            except ValueError:
                bad_numeric += 1
    if bad_numeric:
        warnings.append(f"{bad_numeric} non-numeric / missing vote-share cells")

    if errors:
        print("[build_hts] VALIDATION FAILED:")
        for e in errors:
            print(f"  FAIL  {e}")
        sys.exit(1)

    if warnings:
        for w in warnings:
            print(f"[build_hts] WARN  {w}")

    # ---------------- write CSV ----------------
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(output_rows)
    print(f"[build_hts] wrote {OUT_CSV.relative_to(ROOT.parent)}  ({OUT_CSV.stat().st_size:,} bytes)")

    # ---------------- write XLSX ----------------
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[build_hts] openpyxl not available — skipping XLSX. CSV is ready.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Constituencies"

    # Header
    ws.append(OUTPUT_COLUMNS)
    for col_idx, _ in enumerate(OUTPUT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data
    for r in output_rows:
        ws.append([r[col] for col in OUTPUT_COLUMNS])

    # Column widths (rough autosize)
    width_overrides = {
        "ac_no": 7,
        "constituency": 22,
        "district": 16,
        "region_5way": 11,
        "reservation": 12,
        "2011 Assembly Actual Winner": 32,
        "2014 Lok Sabha AC/Segment Actual Winner": 38,
        "Historical Projection [2011-2014]": 34,
        "2016 Assembly Actual Winner": 16,
        "2021 Assembly Actual Winner": 16,
        "Long-Term Trend [2014-2021]": 22,
        "2021 Assembly Actual Winner For Swing": 22,
        "2024 Lok Sabha Actual Winner": 18,
        "2024 UDF Vote Share (%)": 13,
        "2024 LDF Vote Share (%)": 13,
        "2024 NDA Vote Share (%)": 13,
        "Recent Swing [2021-2024]": 22,
    }
    for col_idx, col in enumerate(OUTPUT_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width_overrides.get(col, 14)
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "B2"

    # Notes sheet
    notes_ws = wb.create_sheet(title="Notes")
    notes_ws.column_dimensions["A"].width = 110
    for line in NOTES_TEXT.split("\n"):
        notes_ws.append([line])
    for cell in notes_ws["A"]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUT_XLSX)
    print(f"[build_hts] wrote {OUT_XLSX.relative_to(ROOT.parent)}  ({OUT_XLSX.stat().st_size:,} bytes)")

    # ---------------- summary ----------------
    print()
    print("[build_hts] Summary:")
    from collections import Counter
    win16 = Counter(r["2016 Assembly Actual Winner"] for r in output_rows)
    win21 = Counter(r["2021 Assembly Actual Winner"] for r in output_rows)
    win24 = Counter(r["2024 Lok Sabha Actual Winner"] for r in output_rows)
    lt = Counter(r["Long-Term Trend [2014-2021]"] for r in output_rows)
    rs = Counter(r["Recent Swing [2021-2024]"] for r in output_rows)
    print(f"  2016 winners: {dict(win16)}")
    print(f"  2021 winners: {dict(win21)}")
    print(f"  2024 winners: {dict(win24)}")
    print(f"  Long-Term Trend distribution: {dict(lt)}")
    print(f"  Recent Swing  distribution:   {dict(rs)}")


if __name__ == "__main__":
    main()
